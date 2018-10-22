# --*-- coding:UTF-8 --*--
import subprocess,os
import xlwt,openpyxl
import configparser
from openpyxl.styles import PatternFill, Border,Side, Alignment,protection,Font,colors
'''
git log refs:https://git-scm.com/book/zh/v2/Git-%E5%9F%BA%E7%A1%80-%E6%9F%A5%E7%9C%8B%E6%8F%90%E4%BA%A4%E5%8E%86%E5%8F%B2
'''
def getTitleType():
    tile_font_style = Font(name='Times New Roman',
                           size=13,
                           bold=True,
                           italic=False,
                           vertAlign=None,
                           underline='none',
                           strike=False,
                           color=colors.BLACK)
    title_fill_style = PatternFill(fill_type='solid',
                                   start_color=colors.GREEN,
                                   end_color=colors.BLACK)
    title_border_style = Border(left=Side(border_style='medium',
                                          color=colors.BLACK),
                                right=Side(border_style='medium',
                                           color=colors.BLACK),
                                top=Side(border_style='medium',
                                         color=colors.BLACK),
                                bottom=Side(border_style='medium',
                                            color=colors.BLACK)
                                )
    title_alig_style = Alignment(horizontal='center',
                                 vertical='center')

    return [tile_font_style,title_fill_style,title_border_style,title_alig_style]

def make_title_hori(wlan_log_sheet,title_str,title_style=None,row_start = 1,colum_start=1,column_end=7,step=1):
    row_index=row_start
    colum_index = colum_start

    if title_style is not None:
        tile_font_style = title_style[0]
        title_fill_style= title_style[1]
        title_border_style = title_style[2]
        title_alig_style = title_style[3]

    for colum in range(colum_start,column_end):
        C = wlan_log_sheet.cell(row=row_index, column=colum_index, value=title_str[colum - 1])
        C.font = tile_font_style
        C.fill = title_fill_style
        C.border = title_border_style
        C.alignment = title_alig_style
        colum_index+=step

####################################################################
#####注意添加的start ID是在此ID后到end ID为止即(startID,endID]
#####
####################################################################
def get_git_log(dir_path, start_id=None, end_id=None,start_date='1970-1-1',end_date='HEAD'):
    if start_id == None:
        cmd = "git log --since=\"%s\" --until=\"%s\" " % (start_date, end_date) + \
              "--name-status --date=iso  --pretty=format:" + \
              _SECTIONSPLITFLAG + \
              "\"%h" + _ITEMSPLITFLAG + \
              "%cn" + _ITEMSPLITFLAG + \
              "%cd" + _ITEMSPLITFLAG + \
              "%s" + _ITEMSPLITFLAG + \
              "\""
    else:
        cmd = "git log %s..%s --since=\"%s\" --until=\"%s\" " % (start_id, end_id, start_date, end_date) + \
              "--name-status --date=iso  --pretty=format:" + \
              _SECTIONSPLITFLAG + \
              "\"%h" + _ITEMSPLITFLAG + \
              "%cn" + _ITEMSPLITFLAG + \
              "%cd" + _ITEMSPLITFLAG + \
              "%s" + _ITEMSPLITFLAG + \
              "\""
    if os.path.isdir(dir_path):
        os.chdir(dir_path)
        SB = subprocess.Popen(cmd,
                              shell=True,
                              stdout=subprocess.PIPE,
                              stderr=subprocess.STDOUT,
                              encoding='utf-8')
        try:
            outs, errors = SB.communicate(timeout=60)
        except subprocess.TimeoutExpired as e:
            SB.kill()
            print(e)
    return outs

def write_to_file(log_sections, conf_parse):
    for section in log_sections:
        section = section.split(_ITEMSPLITFLAG)
        conf_parse[str(section[0])] = {}
        conf_parse[str(section[0])]['commit_id'] = section[0]
        conf_parse[str(section[0])]['commitor'] = section[1]
        conf_parse[str(section[0])]['commit_date'] = section[2]
        conf_parse[str(section[0])]['contents'] = section[3]
        conf_parse[str(section[0])]['diffs'] = section[4]

if __name__ == '__main__':
    _PATH = r'D:\scripts\test'
    _SECTIONSPLITFLAG = 'SECTIONSPLITFLAG'
    _ITEMSPLITFLAG = 'ITERMSPLITFLAG'
    row_index = 1
    colum_index = 1
    sonDirs = os.listdir(_PATH)
    wb  = openpyxl.Workbook()
    wlan_log_sheet = wb.create_sheet('WlanLog',0)
    title_style= getTitleType()
    title_string = ['commitId', 'commitor', 'commit_date', 'contents', 'state', 'path']
    make_title_hori(wlan_log_sheet,title_string,title_style)
    row_index+=1
    conf_parse = configparser.ConfigParser()
    for dir in sonDirs:
        sonDir = os.path.join(_PATH,dir)
        outs = get_git_log(sonDir,start_id='9b6ed7e',end_id="52949bd")
        log_sections = str(outs).split(_SECTIONSPLITFLAG)
        log_sections.pop(0)
        write_to_file(log_sections, conf_parse)
    with open('WlanLog.conf','w') as file_name:
        conf_parse.write(file_name)
        file_name.close()
    conf_parse = configparser.ConfigParser()
    conf_parse.read('WlanLog.conf')
    sections = conf_parse.sections()
    for section in sections:
        wlan_log_sheet.cell(row=row_index, column=colum_index, value=conf_parse[section]['commit_id'])
        wlan_log_sheet.cell(row=row_index, column=colum_index + 1, value=conf_parse[section]['commitor'])
        wlan_log_sheet.cell(row=row_index, column=colum_index + 2, value=conf_parse[section]['commit_date'])
        wlan_log_sheet.cell(row=row_index, column=colum_index + 3, value=conf_parse[section]['contents'])
        diffs = conf_parse[section]['diffs'].split()
        wlan_log_sheet.cell(row=row_index, column=colum_index + 4, value=diffs[0])
        wlan_log_sheet.cell(row=row_index, column=colum_index + 5, value=diffs[1])
        row_index+=1
    wb.save('Wlan Log.xlsx')

